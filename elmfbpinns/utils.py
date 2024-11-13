import numpy as np
import pandas as pd
import jax.numpy as jnp
import openpyxl

# Helper function to display large matrices
def display_matrix(M, num_columns=None):
    df = pd.DataFrame(M)
    if num_columns is None or num_columns > len(df.columns):
        num_columns = len(df.columns)
    print(df.iloc[:, :num_columns])


# Helper function to display window intervals
def display_windows(xmins, xmaxs):
    data = {
        "Window #": [f"Window {i+1}" for i in range(len(xmins))],
        "xmins": xmins,
        "xmaxs": xmaxs,
    }
    df = pd.DataFrame(data)
    print(df)
    
#loss functions
def calc_l1_loss(u, f_x):
    loss = jnp.mean(jnp.abs(f_x - u)).item()
    print(f"Loss Value: {loss:.2e}")
    return loss


def calc_normalized_l1_loss(a, M, b):
    residual = M @ a - b
    l1_loss = jnp.mean(jnp.abs(residual))
    b_std = b.std()
    normalized_l1_loss = l1_loss / b_std
    print(f"Normalized L1 Loss Value: {normalized_l1_loss:.2e}")

    return normalized_l1_loss

#Save and load to and from excel files
def save_results_to_excel(
    Ms, B, a, us, losses, xs, fs, xmins, xmaxs, elapsed_time, lhs_condition, filename
):
    import openpyxl
    from openpyxl import Workbook

    # Create a new workbook
    workbook = Workbook()

    # Add matrices Ms to separate sheets
    for i, M in enumerate(Ms):
        sheet_name = f"Matrix_M_{i+1}"
        worksheet = workbook.create_sheet(sheet_name)
        M = jnp.asarray(M)  # Ensure it's a NumPy array if it's a JAX array
        for row in range(M.shape[0]):
            for col in range(M.shape[1]):
                worksheet.cell(row=row + 1, column=col + 1, value=float(M[row, col]))

    # Add matrix B to a separate sheet
    for i, B_matrix in enumerate(B):
        sheet_name = f"Matrix_B_{i+1}"
        worksheet = workbook.create_sheet(sheet_name)
        # Ensure it's a NumPy array if it's a JAX array
        B_matrix = jnp.asarray(B_matrix)
        for row in range(B_matrix.shape[0]):
            for col in range(B_matrix.shape[1]):
                worksheet.cell(
                    row=row + 1, column=col + 1, value=float(B_matrix[row, col])
                )

    # Add vector a to a sheet
    worksheet = workbook.create_sheet("Vector_a")
    a = jnp.asarray(a)
    for i in range(a.shape[0]):
        worksheet.cell(row=i + 1, column=1, value=float(a[i]))

    # Add us to a sheet
    for i, u in enumerate(us):
        sheet_name = f"U_{i+1}"
        worksheet = workbook.create_sheet(sheet_name)
        u = jnp.asarray(u)
        for j in range(u.shape[0]):
            worksheet.cell(row=j + 1, column=1, value=float(u[j]))

    # Add losses to a sheet
    worksheet = workbook.create_sheet("Losses")
    for i, loss in enumerate(losses):
        worksheet.cell(row=i + 1, column=1, value=float(loss))

    # Add xs to a sheet
    for i, x in enumerate(xs):
        sheet_name = f"X_{i+1}"
        worksheet = workbook.create_sheet(sheet_name)
        x = jnp.asarray(x)
        for j in range(x.shape[0]):
            worksheet.cell(row=j + 1, column=1, value=float(x[j]))

    # Add fs to a sheet
    for i, f in enumerate(fs):
        sheet_name = f"F_{i+1}"
        worksheet = workbook.create_sheet(sheet_name)
        f = jnp.asarray(f)
        for j in range(f.shape[0]):
            worksheet.cell(row=j + 1, column=1, value=float(f[j]))

    # Add xmins and xmaxs to a sheet
    worksheet = workbook.create_sheet("Intervals")
    for i in range(len(xmins)):
        worksheet.cell(row=i + 1, column=1, value=float(xmins[i]))
        worksheet.cell(row=i + 1, column=2, value=float(xmaxs[i]))

    # Add elapsed_time to a sheet
    worksheet = workbook.create_sheet("Elapsed_Time")
    worksheet.cell(row=1, column=1, value=float(elapsed_time))

    # Add condition numbers to a sheet
    worksheet = workbook.create_sheet("Condition_Numbers")
    worksheet.cell(row=1, column=1, value=float(lhs_condition))

    # Remove the default sheet created with the workbook
    default_sheet = workbook["Sheet"]
    workbook.remove(default_sheet)

    # Save the workbook
    workbook.save(filename)


def load_results_from_excel(filename):
    import openpyxl
    from openpyxl import load_workbook
    import numpy as np

    # Load the workbook
    workbook = load_workbook(filename)

    # Load matrices Ms from sheets
    Ms = []
    i = 1
    while f"Matrix_M_{i}" in workbook.sheetnames:
        sheet = workbook[f"Matrix_M_{i}"]
        M = np.array(
            [
                [
                    sheet.cell(row=row + 1, column=col + 1).value
                    for col in range(sheet.max_column)
                ]
                for row in range(sheet.max_row)
            ]
        )
        Ms.append(M)
        i += 1

    # Load matrix B from sheets
    Bs = []
    i = 1
    while f"Matrix_B_{i}" in workbook.sheetnames:
        sheet = workbook[f"Matrix_B_{i}"]
        B = np.array(
            [
                [
                    sheet.cell(row=row + 1, column=col + 1).value
                    for col in range(sheet.max_column)
                ]
                for row in range(sheet.max_row)
            ]
        )
        Bs.append(B)
        i += 1

    # Load vector a
    sheet = workbook["Vector_a"]
    a = np.array([sheet.cell(row=i + 1, column=1).value for i in range(sheet.max_row)])

    # Load us
    us = []
    i = 1
    while f"U_{i}" in workbook.sheetnames:
        sheet = workbook[f"U_{i}"]
        u = np.array(
            [sheet.cell(row=j + 1, column=1).value for j in range(sheet.max_row)]
        )
        us.append(u)
        i += 1

    # Load losses
    sheet = workbook["Losses"]
    losses = np.array(
        [sheet.cell(row=i + 1, column=1).value for i in range(sheet.max_row)]
    )

    # Load xs
    xs = []
    i = 1
    while f"X_{i}" in workbook.sheetnames:
        sheet = workbook[f"X_{i}"]
        x = np.array(
            [sheet.cell(row=j + 1, column=1).value for j in range(sheet.max_row)]
        )
        xs.append(x)
        i += 1

    # Load fs
    fs = []
    i = 1
    while f"F_{i}" in workbook.sheetnames:
        sheet = workbook[f"F_{i}"]
        f = np.array(
            [sheet.cell(row=j + 1, column=1).value for j in range(sheet.max_row)]
        )
        fs.append(f)
        i += 1

    # Load xmins and xmaxs
    sheet = workbook["Intervals"]
    xmins = np.array(
        [sheet.cell(row=i + 1, column=1).value for i in range(sheet.max_row)]
    )
    xmaxs = np.array(
        [sheet.cell(row=i + 1, column=2).value for i in range(sheet.max_row)]
    )

    # Load elapsed_time
    sheet = workbook["Elapsed_Time"]
    elapsed_time = sheet.cell(row=1, column=1).value

    # Load condition numbers
    sheet = workbook["Condition_Numbers"]
    lhs_condition = sheet.cell(row=1, column=1).value

    return Ms, Bs, a, us, losses, xs, fs, xmins, xmaxs, elapsed_time, lhs_condition


def extract_FBPINN_PINN_results(file_path):
    # Read the Excel file
    xls = pd.ExcelFile(file_path)

    # Extract data from the "results" sheet
    results_df = pd.read_excel(xls, sheet_name="results")
    u_pinn = results_df["u_pinn"].values
    u_fbpinn = results_df["u_fbpinn"].values

    # Extract data from the "time_loss" sheet
    time_loss_df = pd.read_excel(xls, sheet_name="time_loss")
    fbpinn_time_taken = time_loss_df["FBPINN_time_taken"].values
    fbpinn_l1_losses = time_loss_df["FBPINN_l1_losses"].values
    fbpinn_norm_l1_losses = time_loss_df["FBPINN_l1_norm_losses"].values
    pinn_time_taken = time_loss_df["PINN_time_taken"].values
    pinn_l1_losses = time_loss_df["PINN_l1_losses"].values
    pinn_norm_l1_losses = time_loss_df["PINN_l1_norm_losses"].values

    # Extract data from the "metrics" sheet
    metrics_df = pd.read_excel(xls, sheet_name="metrics")
    metrics = metrics_df.set_index("Metric")["Value"].to_dict()
    final_loss_value_pinn = metrics["final_loss_value_pinn"]
    final_loss_value_fbpinn = metrics["final_loss_value_fbpinn"]
    final_time_taken_pinn = metrics["final_time_taken_pinn"]
    final_time_taken_fbpinn = metrics["final_time_taken_fbpinn"]

    # Return all data in a dictionary for easy access
    data = {
        "u_pinn": u_pinn,
        "u_fbpinn": u_fbpinn,
        "fbpinn_time_taken": fbpinn_time_taken,
        "fbpinn_l1_losses": fbpinn_l1_losses,
        "fbpinn_l1_norm_losses": fbpinn_norm_l1_losses,
        "pinn_time_taken": pinn_time_taken,
        "pinn_l1_losses": pinn_l1_losses,
        "pinn_l1_norm_losses": pinn_norm_l1_losses,
        "final_loss_value_pinn": final_loss_value_pinn,
        "final_loss_value_fbpinn": final_loss_value_fbpinn,
        "final_time_taken_pinn": final_time_taken_pinn,
        "final_time_taken_fbpinn": final_time_taken_fbpinn,
    }

    return data

# Extract the lhs_condition and network size from the Excel files
def extract_lhs_cond_network_size(activation, C, R, n_test, J_min, J_max):
    # Initialize empty lists to store the results
    lhs_conditions = []
    network_sizes = []
    network_sizes_no_n = []

    # Loop over the file numbers from J_min to J_max
    for i in range(J_min, J_max + 1):
        # Construct the filename based on the pattern
        filename = f"{activation}_J{i}_C{C}_ntest{n_test}_R{R}.xlsx"

        # Load the workbook
        workbook = openpyxl.load_workbook(filename, data_only=True)

        # Select the 'Condition_Numbers' sheet
        worksheet = workbook["Condition_Numbers"]

        # Get the lhs_condition number from the first cell (A1)
        lhs_condition = worksheet.cell(row=1, column=1).value

        # Append the lhs_condition number to the list
        lhs_conditions.append(lhs_condition)
        network_sizes.append(150 * 32 * i)
        network_sizes_no_n.append(32 * i)

    # Convert the lists to numpy arrays for easier manipulation
    lhs_conditions_array = np.array(lhs_conditions)
    network_sizes_array = np.array(network_sizes)
    network_sizes_no_n_array = np.array(network_sizes_no_n)

    return lhs_conditions_array, network_sizes_array, network_sizes_no_n_array